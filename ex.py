import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import os

patients = []  
patient_id_counter = 601  
doctor_name = "John Doe"  
excel_file = "patient_data.xlsx"  

def start_application():
    start_button.pack_forget()  
    create_main_window()

def create_main_window():

    background_image = Image.open("E:/my self/project/pre-care for diagones/b5.webp")
                
    background_photo = ImageTk.PhotoImage(background_image)
    background_label = tk.Label(root, image=background_photo)
    background_label.image = background_photo
    background_label.place(relwidth=1, relheight=1)

    
    title_label = tk.Label(root, text="   PRE-CARE FOR DIAGNOSIS   ", font=("Times New Roman ", 25, "bold"), fg="black",bg="blue")
    title_label.place(relx=0.5, rely=0.05, anchor="center")  

    
    name_label = tk.Label(root, text="Name:", font=("Helvetica", 20, "bold"), bg="lightblue")
    name_label.place(relx=0.3, rely=0.15, anchor="center")
    global name_entry
    name_entry = tk.Entry(root, font=("Helvetica", 20, "bold"))
    name_entry.place(relx=0.5, rely=0.15, anchor="center")

    
    age_label = tk.Label(root, text="Age:", font=("Helvetica", 20, "bold"), bg="lightblue")
    age_label.place(relx=0.3, rely=0.25, anchor="center")
    global age_entry
    age_entry = tk.Entry(root, font=("Helvetica", 20, "bold"))
    age_entry.place(relx=0.5, rely=0.25, anchor="center")

    
    symptoms_label = tk.Label(root, text="Symptoms:", font=("Helvetica", 20, "bold"), bg="lightblue")
    symptoms_label.place(relx=0.3, rely=0.35, anchor="center")
    global symptoms_entry
    symptoms_entry = tk.Entry(root, font=("Helvetica", 20, "bold"))
    symptoms_entry.place(relx=0.5, rely=0.35, anchor="center")

    add_button = tk.Button(root, text="Add Patient", command=add_patient, font=("Helvetica", 20, "bold"), bg="green", fg="white")
    add_button.place(relx=0.4, rely=0.45, anchor="center")

    view_button = tk.Button(root, text="View Patients", command=view_patients, font=("Helvetica", 20, "bold"), bg="orange", fg="white")
    view_button.place(relx=0.6, rely=0.45, anchor="center")

    diagnose_button = tk.Button(root, text="Diagnose", command=diagnose_patient, font=("Helvetica", 20, "bold"), bg="purple", fg="white")
    diagnose_button.place(relx=0.4, rely=0.55, anchor="center")

    export_button = tk.Button(root, text="Save to Excel", command=save_to_excel, font=("Helvetica", 20, "bold"), bg="blue", fg="white")
    export_button.place(relx=0.6, rely=0.55, anchor="center")

    
    global patient_info_label
    patient_info_label = tk.Label(root, text="", justify="left", font=("Helvetica", 15, "bold"), bg="lightblue")
    patient_info_label.place(relx=0.5, rely=0.93, anchor="center")

    global diagnosis_label
    diagnosis_label = tk.Label(root, text="", justify="left", font=("Helvetica", 15, "bold"), bg="lightblue")
    diagnosis_label.place(relx=0.5, rely=0.75, anchor="center")

def add_patient():
    global patient_id_counter  

    name = name_entry.get()
    age = age_entry.get()
    symptoms = symptoms_entry.get()

    if not name or not age or not symptoms:
        messagebox.showerror("Input Error", "All fields are required.")
        return

    try:
        age = int(age)  
    except ValueError:
        messagebox.showerror("Input Error", "Age must be a valid number.")
        return

    diagnosis = get_diagnosis(symptoms.lower())

    patient = {
        "id": patient_id_counter,  
        "name": name,
        "age": age,
        "symptoms": symptoms,
        "diagnosis": diagnosis
    }

    patients.append(patient)
    patient_id_counter += 1  
    clear_entries()
    messagebox.showinfo("Success", f"Patient {patient['id']} added successfully.")
    diagnosis_label.config(text=diagnosis) 

def view_patients():
    if not patients:
        patient_info_label.config(text="No patients found.")
    else :
        patient_info = "\n".join([f"ID: {p['id']}, Name: {p['name']}, Age: {p['age']}, Symptoms: {p['symptoms']}, Diagnosis: {p['diagnosis']}" for p in patients])
        patient_info_label.config(text=patient_info)

def diagnose_patient():
    symptoms = symptoms_entry.get().lower()
    if not symptoms:
        messagebox.showerror("Input Error", "Symptoms field is required.")
        return

    diagnosis = get_diagnosis(symptoms)
    diagnosis_label.config(text=diagnosis)

def get_diagnosis(symptoms):
    
    diagnoses = {
        "fever": ("heat exhaustion", "dolo 650, paracetamol", "Rest and hydrate"),
        "stomach pain": ("indigestion, gastritis", "antacids, ENO", "Eat easily digested meals, avoid junk food"),
        "back pain": ("muscle strain, disc issues", "pain relievers, tiger balm", "Maintain good posture, exercise regularly"),
        "runny nose": ("cold, sinus", "coldact, benadryl", "Drink hot water, stay hydrated"),
        "cough": ("cold, sinus", "coldact, benadryl", "Drink hot water, stay hydrated"),
        "diabetes mellitus": ("frequent urination, blurred vision", "insulin, oral hypoglycemic agents", "Monitor blood glucose levels"),
        "hypertension": ("headache, nosebleeds", "antihypertensive medications", "Maintain a healthy weight, low-sodium diet"),
        "asthma":("shortness of breath, chest tightness", "inhaled corticosteroids, bronchodilators", "Avoid triggers"),
        "chicken pox": ("itchy rash, fever", "antihistamines, calamine lotion", "Varicella vaccine, avoid close contact with infected individuals"),
        "body pain":("body pains , reducing fever ","acetaminophen,tylenol","Regular health check-up"),
        "bp":("shortness of breath , chest pain "," ACE-inhibitor , Beta Blocker "," Monitor blood pressure regularly"),
        "eye dryness":("eye redness , watery eyes ","Artificial tears , prescription eye drops "," stay hydrated , take break"),
        "platelets":("nose bleeds , fatigue ","corticosteroids , intravenous immunoglobulin(IVIG)","Stay Hydrated and Eat a Balanced Diet"),
        "dizziness":(" unsteadiness ","antihistamines , anticholinergics ","stay hydrated , manage stress"),
        "knee pains":(" instability , stiffness ","acetaminophen , Ibuprofen","exercise regularly , use proper footwear"),
        "skin rashes":("heat rash,infection ", "Hydrocortisone cream,Antibacterial ointments "," fragrance-free soaps, lotions"),
        "snake bite":("Cobras,Coral Snakes","acetaminophen,Antibiotics","Be Aware of Your Surroundings"),
        "food poisoning":("Bacterial Causes,Viral Causes","Antibiotics,Rehydration","Traveling Safely,Proper Food Storage"),
        "fits":("Alcohol Withdrawal Seizures","Phenytoin,Valproate","Healthy Lifestyle,avoid triggers"),



        }

    for key, (cause, medicines, prescription) in diagnoses.items():
        if key in symptoms:
            return f"\npossible causes: {cause} ,     Medicines: {medicines},      Prescription: {prescription}"
    return "Diagnosis: If you join emergency hospital immediately."

def clear_entries():
    name_entry.delete(0, tk.END)
    age_entry.delete(0, tk.END)
    symptoms_entry.delete(0, tk.END)

def save_to_excel():    
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Patient Data"
        headers = ["ID", "Name", "Age", "Symptoms", "Diagnosis"]
        sheet.append(headers)

    sheet = workbook.active

    for patient in patients:
        row = [patient["id"], patient["name"], patient["age"], patient["symptoms"], patient["diagnosis"]]
        sheet.append(row)
    try:
        workbook.save(excel_file)
        messagebox.showinfo("Success", f"Patient data saved successfully to '{excel_file}'.")
    except Exception as e:
        messagebox.showerror("Save Error", f"Failed to save data: {e}")

root = tk.Tk()
root.geometry("1000x1000")  
root.title("PRE-CARE FOR DIAGNOSIS")

background_image = Image.open("E:/my self/project/pre-care for diagones/36.png")  
background_photo = ImageTk.PhotoImage(background_image)
    
background_label =  tk.Label(root, image=background_photo)
background_label.image = background_photo  
background_label.place(relwidth=1, relheight=1)
title_label = tk.Label(root, text="  PRE-CARE FOR DIAGNOSIS  ", font=("Helvetica", 50, "bold"), fg="black", bg="darkred")
title_label.place(relx=0.5, rely=0.11, anchor="center") 

start_button = tk.Button(root, text=" Click Here ", command=start_application, font=("Helvetica", 25, "bold"), bg="blue", fg="black")
start_button.place(relx=0.8, rely=0.8, anchor="center")

root.mainloop()


